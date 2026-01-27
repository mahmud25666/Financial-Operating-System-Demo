import streamlit as st
import pandas as pd
import os
import time
import base64
import plotly.express as px
from datetime import datetime
from openpyxl.styles import PatternFill, Font

# --- CONFIGURATION ---
st.set_page_config(page_title="Glafit Empire Finance V5", layout="wide", page_icon="🏢")
FILE = 'Finance_Master_V5.xlsx'
VAULT = 'Master_Vault'

import pdfplumber
import os
import re
from datetime import datetime
import dateparser

def clean_text(text):
    if not text: return ""
    text = text.replace('：', ':').replace('\xa0', ' ')
    # Merge hyphenated words across lines (e.g., "Con-\ntract" -> "Contract")
    text = re.sub(r'(\w+)-\n(\w+)', r'\1\2', text)
    return text

def find_amount_in_text(text):
    """Finds the largest number that looks like currency"""
    # Supports JPY format 10,025,000 and USD 10,000.00
    pattern = r"(?:[A-Z]{3}|[\$£€¥])?\s?([\d,]+\.?\d{0,2})"
    matches = re.findall(pattern, text)
    valid_amts = []
    for m in matches:
        try:
            val = float(re.sub(r'[^\d.]', '', m))
            if val > 0: valid_amts.append(val)
        except: pass
    return max(valid_amts) if valid_amts else 0.0

def find_amount_in_block(text_block):
    """(Alias for compatibility) Finds amount in a specific block"""
    return find_amount_in_text(text_block)

def extract_dynamic_date(text, default=None):
    # Matches 2025/12/05, 05-Dec-2025, 12th Jan 2025 etc.
    regex = r'(?:Date|Dated|On)?\s*[:.]?\s*(\d{4}[./-]\d{1,2}[./-]\d{1,2}|\d{1,2}[./-]\d{1,2}[./-]\d{2,4})'
    match = re.search(regex, text, re.IGNORECASE)
    if match:
        try:
            dt = dateparser.parse(match.group(1))
            if dt: return dt.date()
        except: pass
    return default or datetime.now().date()

def parse_multi_sow_agreement(pdf_path):
    """
    ROBUST PARSER: Scans for multiple 'Scope of Work' sections.
    """
    extracted_sows = []
    global_date = datetime.now().date()
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = "\n".join([p.extract_text() for p in pdf.pages if p.extract_text()])
            text = clean_text(full_text)
            
            # 1. Find Global Document Date
            global_date = extract_dynamic_date(text[:1000])

            # 2. ROBUST REGEX for SOW
            # Matches: "Scope of Work (SOW):" OR "Scope of Work:" OR "SOW" followed by newline
            # (?:\s*\(SOW\))?  --> Optional (SOW)
            # \s*[:\n]         --> Matches a Colon OR a New Line (The fix!)
            sow_pattern = re.compile(r"(?:Scope of Work|SOW)(?:\s*\(SOW\))?\s*[:\n]", re.IGNORECASE)
            
            matches = list(sow_pattern.finditer(text))
            
            if not matches:
                # Fallback: Treat whole doc as one SOW
                amt = find_amount_in_text(text)
                return [{'name': 'General Agreement', 'amount': amt, 'date': global_date}]

            # 3. Iterate matches
            for i, match in enumerate(matches):
                start_idx = match.start()
                end_idx = match.end()
                
                # A. Find Name (Look backwards)
                preceding_text = text[max(0, start_idx-200):start_idx]
                lines = [l.strip() for l in preceding_text.split('\n') if l.strip()]
                
                # The line immediately before "Scope of Work" is usually the Title
                sow_name = lines[-1] if lines else f"Project Section {i+1}"
                # Clean bullet points "1. Project..." -> "Project..."
                sow_name = re.sub(r'^[\d.\-\)]+\s*', '', sow_name) 

                # B. Find Content (Look forwards until next match)
                next_start = matches[i+1].start() if i+1 < len(matches) else len(text)
                block_text = text[end_idx:next_start]
                
                # C. Extract Data
                amount = find_amount_in_text(block_text)
                section_date = extract_dynamic_date(block_text, default=global_date)
                
                extracted_sows.append({
                    'name': sow_name,
                    'amount': amount,
                    'date': section_date
                })
                
    except Exception as e:
        print(f"Error parsing SOW: {e}")
        return []

    return extracted_sows

def parse_invoice_v2(pdf_path):
    """Advanced Invoice Parser with Table Support"""
    inv_data = {
        'no': "DRAFT", 
        'date': datetime.now().date(), 
        'total': 0.0,
        'items': [] 
    }
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            first_page_text = clean_text(pdf.pages[0].extract_text())
            
            # Header Info
            no_match = re.search(r"(?:Invoice\s*No\.?|No\.?)\s*[:.]?\s*([A-Z0-9\-\/]{5,})", first_page_text, re.IGNORECASE)
            if no_match: inv_data['no'] = no_match.group(1).strip()
            
            inv_data['date'] = extract_dynamic_date(first_page_text)
            inv_data['total'] = find_amount_in_text(first_page_text)

            # Table Scan
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        cleaned_row = [clean_text(str(cell)) if cell else "" for cell in row]
                        
                        # Find Description & Amount columns
                        desc_candidates = [c for c in cleaned_row if len(c) > 5 and not re.match(r'^[\d,.]+$', c)]
                        amt_candidates = [c for c in cleaned_row if re.search(r'[\d,]+\.?\d*', c)]
                        
                        if desc_candidates and amt_candidates:
                            desc = desc_candidates[0].replace('\n', ' ')
                            if "Description" in desc or "Item" in desc: continue
                            
                            try:
                                amt_txt = amt_candidates[-1]
                                amt = float(re.sub(r'[^\d.]', '', amt_txt))
                                if amt > 0:
                                    inv_data['items'].append({'desc': desc, 'amount': amt})
                            except: pass

    except Exception as e:
        print(f"Error parsing PDF: {e}")

    if not inv_data['items']:
        inv_data['items'].append({'desc': 'General Services', 'amount': inv_data['total']})
        
    return inv_data

def parse_payment(pdf_path):
    """Simple Payment Parser"""
    inv_ref = None
    amt = 0.0
    p_date = datetime.now().date()
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = clean_text(pdf.pages[0].extract_text() or "")
            amt = find_amount_in_text(text)
            p_date = extract_dynamic_date(text)
            
            match = re.search(r"(INV-[A-Z0-9\-]+)", text)
            if match: inv_ref = match.group(1).strip()
    except: pass
    
    return inv_ref, amt, p_date

if not os.path.exists(VAULT):
    os.makedirs(VAULT)

# --- DATABASE SCHEMA ---
COLS_QT = ['Quote_ID', 'Date', 'Business', 'Project_Name', 'Total_Value', 'Agreement_File', 'Status']
COLS_INV = ['Invoice_No', 'Quote_Ref', 'Date', 'Business', 'Split_Amount', 'Description', 'Invoice_File', 'Declaration_File']
# ✅ NEW: Quote_Ref added so payments can be allocated per quotation (important when one invoice maps to multiple quotes)
COLS_PAY = ['Payment_ID', 'Parent_Payment_ID', 'Invoice_Ref', 'Quote_Ref', 'Date', 'Amount', 'Proof_File', 'Form_C_File', 'Payment_Decl_File']


# --- HELPER FUNCTIONS ---
def load_db():
    try:
        df_q = pd.read_excel(FILE, sheet_name='Quotations')
        df_i = pd.read_excel(FILE, sheet_name='Invoices')
        df_p = pd.read_excel(FILE, sheet_name='Payments')

        # Ensure Columns Exist
        for c in COLS_QT:
            if c not in df_q.columns:
                df_q[c] = ""
        for c in COLS_INV:
            if c not in df_i.columns:
                df_i[c] = ""
        for c in COLS_PAY:
            if c not in df_p.columns:
                df_p[c] = ""

        # Force Numeric Types
        df_q['Total_Value'] = pd.to_numeric(df_q['Total_Value'], errors='coerce').fillna(0.0)
        df_i['Split_Amount'] = pd.to_numeric(df_i['Split_Amount'], errors='coerce').fillna(0.0)
        df_p['Amount'] = pd.to_numeric(df_p['Amount'], errors='coerce').fillna(0.0)

        # Force String Types
        df_q['Quote_ID'] = df_q['Quote_ID'].astype(str)
        df_i['Invoice_No'] = df_i['Invoice_No'].astype(str)
        df_i['Quote_Ref'] = df_i['Quote_Ref'].astype(str)
        df_p['Invoice_Ref'] = df_p['Invoice_Ref'].astype(str)
        df_p['Quote_Ref'] = df_p['Quote_Ref'].fillna("").astype(str)
        df_p['Payment_ID'] = df_p['Payment_ID'].astype(str)
        df_p['Parent_Payment_ID'] = df_p['Parent_Payment_ID'].fillna("").astype(str)

        # Force File Paths
        for col in ['Agreement_File']:
            if col in df_q.columns:
                df_q[col] = df_q[col].fillna("None").astype(str)
        for col in ['Invoice_File', 'Declaration_File']:
            if col in df_i.columns:
                df_i[col] = df_i[col].fillna("None").astype(str)
        for col in ['Proof_File', 'Form_C_File', 'Payment_Decl_File']:
            if col in df_p.columns:
                df_p[col] = df_p[col].fillna("None").astype(str)

        return df_q, df_i, df_p

    except Exception:
        return (
            pd.DataFrame(columns=COLS_QT),
            pd.DataFrame(columns=COLS_INV),
            pd.DataFrame(columns=COLS_PAY)
        )


def safe_copy(file_obj, folder_path, file_name):
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    if not file_obj:
        return "None"
    final_path = os.path.join(folder_path, file_name)
    with open(final_path, "wb") as f:
        f.write(file_obj.getbuffer())
    return file_name


def display_pdf(file_obj_or_path):
    data = None
    try:
        if hasattr(file_obj_or_path, 'getvalue'):
            data = file_obj_or_path.getvalue()
        elif isinstance(file_obj_or_path, str):
            if file_obj_or_path == "None" or not os.path.exists(file_obj_or_path):
                st.warning("⚠️ File not found.")
                return
            with open(file_obj_or_path, "rb") as f:
                data = f.read()

        if data:
            base64_pdf = base64.b64encode(data).decode('utf-8')
            pdf_display = f'<iframe src="data:application/pdf;base64,{base64_pdf}" width="100%" height="600" type="application/pdf"></iframe>'
            st.markdown(pdf_display, unsafe_allow_html=True)
    except Exception:
        pass


def _payments_for_invoice_quote(df_p, inv_no: str, quote_id: str):
    """
    If payments are allocated (Quote_Ref present), use it.
    Otherwise fallback to invoice-only payments.
    """
    df = df_p[df_p['Invoice_Ref'].astype(str) == str(inv_no)].copy()
    if 'Quote_Ref' in df.columns and df['Quote_Ref'].astype(str).str.len().sum() > 0:
        return df[df['Quote_Ref'].astype(str) == str(quote_id)]
    return df


def generate_ledger_view(curr_biz, df_q, df_i, df_p):
    rows = []

    qs = df_q[df_q['Business'] == curr_biz].copy()
    invs = df_i[df_i['Business'] == curr_biz].copy()
    pays = df_p.copy()

    grand_billed = 0.0
    grand_collected = 0.0

    for _, q in qs.iterrows():
        qid = str(q['Quote_ID'])
        q_val = float(q['Total_Value'])

        q_invs = invs[invs['Quote_Ref'] == qid]
        q_billed = float(q_invs['Split_Amount'].sum())

        # ✅ Quote header Remaining Quoted Balance = Total Quote - Total Billed
        q_unbilled = q_val - q_billed

        # ✅ Collected against this Quote (works even if invoice spans multiple quotes)
        q_collected = 0.0
        for inv_no in q_invs['Invoice_No'].astype(str).unique().tolist():
            q_collected += float(_payments_for_invoice_quote(pays, inv_no, qid)['Amount'].sum())

        # 1) QUOTE HEADER
        rows.append({
            'Type': 'QUOTE', 'Ref': qid, 'Date': q['Date'],
            'Description': f"📂 PROJECT: {q['Project_Name']}",
            'Debit': q_val,                 # Contract Value
            'Credit': q_collected,          # Cash In (Allocated to this Quote)
            'Balance': q_unbilled,          # Remaining to bill
            'Status': "⏳" if q_unbilled > 1.0 else "✅"
        })

        # 2) INVOICE LOOP (per quote)
        for _, i in q_invs.iterrows():
            inv_no = str(i['Invoice_No'])
            inv_amt = float(i['Split_Amount'])

            rows.append({
                'Type': 'INVOICE', 'Ref': inv_no, 'Date': i['Date'],
                'Description': f"  ↳ 🧾 Inv: {i['Description']}",
                'Debit': inv_amt, 'Credit': 0, 'Balance': 0, 'Status': ''
            })

            # 3) PAYMENT LOOP (allocated by quote+invoice)
            my_pays = _payments_for_invoice_quote(pays, inv_no, qid)
            inv_collected = 0.0

            for _, p in my_pays.iterrows():
                p_amt = float(p['Amount'])
                inv_collected += p_amt

                icons = []
                if str(p.get('Proof_File', "None")) != "None":
                    icons.append("🏦")
                if str(p.get('Form_C_File', "None")) != "None":
                    icons.append("📄")
                if str(p.get('Payment_Decl_File', "None")) != "None":
                    icons.append("📝")

                rows.append({
                    'Type': 'PAYMENT', 'Ref': str(p['Payment_ID']), 'Date': p['Date'],
                    'Description': f"    ↳ 💰 Payment Received {' '.join(icons)}",
                    'Debit': 0, 'Credit': p_amt, 'Balance': 0, 'Status': ''
                })

            # ✅ Invoice status line with % covered directly under invoice (strict)
            inv_bal = inv_amt - inv_collected
            pct = (inv_collected / inv_amt * 100.0) if inv_amt > 0 else 0.0
            status_icon = "✅" if inv_bal < 1.0 else "🔴"

            rows.append({
                'Type': 'SUB_SUM', 'Ref': '', 'Date': '',
                'Description': f"    👉 Status: {pct:.1f}% Cleared (Due: {inv_bal:,.0f})",
                'Debit': 0, 'Credit': 0, 'Balance': inv_bal, 'Status': status_icon
            })

        # 4) QUOTE SUMMARY ROW: show Unbilled vs Unpaid
        q_unpaid = q_billed - q_collected
        rows.append({
            'Type': 'SUMMARY', 'Ref': 'TOTAL', 'Date': '',
            'Description': f"📊 PROJECT TOTALS | Unbilled: {q_unbilled:,.0f} | Unpaid: {q_unpaid:,.0f} | Billed: {q_billed:,.0f}",
            'Debit': q_billed, 'Credit': q_collected, 'Balance': q_unpaid,
            'Status': "✅" if q_unpaid < 1.0 else "🔴"
        })
        rows.append({'Type': 'SPACE'})

        grand_billed += q_billed
        grand_collected += q_collected

    # GRAND TOTAL
    grand_outstanding = grand_billed - grand_collected
    g_pct = (grand_collected / grand_billed * 100.0) if grand_billed > 0 else 0.0
    g_status = "🟢" if grand_outstanding < 1.0 else "🔴"

    rows.append({
        'Type': 'GRAND', 'Ref': 'ALL', 'Date': datetime.today(),
        'Description': f"BUSINESS GRAND TOTAL ({g_pct:.1f}% Collected)",
        'Debit': grand_billed, 'Credit': grand_collected, 'Balance': grand_outstanding, 'Status': g_status
    })

    return pd.DataFrame(rows)


def save_db(df_q, df_i, df_p, curr_biz=None):
    with pd.ExcelWriter(FILE, engine='openpyxl') as writer:
        df_q.to_excel(writer, sheet_name='Quotations', index=False)
        df_i.to_excel(writer, sheet_name='Invoices', index=False)
        df_p.to_excel(writer, sheet_name='Payments', index=False)

        if curr_biz:
            ledger_df = generate_ledger_view(curr_biz, df_q, df_i, df_p)
            ledger_df.to_excel(writer, sheet_name='Master_Ledger_View', index=False)

            ws = writer.book['Master_Ledger_View']
            fill_q = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            fill_inv = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            fill_pay = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            fill_sum = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            fill_grand = PatternFill(start_color="212121", end_color="212121", fill_type="solid")
            font_grand = Font(color="FFFFFF", bold=True)
            red_font = Font(color="FF0000", bold=True)
            green_font = Font(color="008000", bold=True)

            for row in ws.iter_rows(min_row=2):
                rtype = row[0].value
                status = row[7].value

                if rtype == 'QUOTE':
                    for cell in row:
                        cell.fill = fill_q
                        cell.font = Font(bold=True)
                elif rtype == 'INVOICE':
                    for cell in row:
                        cell.fill = fill_inv
                elif rtype == 'PAYMENT':
                    for cell in row:
                        cell.fill = fill_pay
                elif rtype == 'SUMMARY':
                    for cell in row:
                        cell.fill = fill_sum
                        cell.font = Font(bold=True)
                elif rtype == 'GRAND':
                    for cell in row:
                        cell.fill = fill_grand
                        cell.font = font_grand

                if status == '🔴':
                    row[7].font = red_font
                elif status == '✅' or status == '🟢':
                    row[7].font = green_font


# --- APP START ---
df_q, df_i, df_p = load_db()

st.sidebar.title("🏢 Glafit Finance")
all_biz = list(set(df_q['Business'].unique().tolist() + ["Glafit_Main"]))
curr_biz = st.sidebar.selectbox("Select Business Unit", all_biz + ["+ New Business"])
if curr_biz == "+ New Business":
    curr_biz = st.sidebar.text_input("New Business Name", "New_Unit_Name")

st.title(f"🚀 Operations: {curr_biz}")
tab0, tab1, tab2, tab3, tab4 = st.tabs(["📈 Dashboard", "1️⃣ Quotations", "2️⃣ Invoices", "3️⃣ Payments", "📊 Master Ledger"])


# --- TAB 0: DASHBOARD ---
with tab0:
    st.write("### 📊 Executive Overview")

    qs = df_q[df_q['Business'] == curr_biz]
    invs = df_i[df_i['Business'] == curr_biz]
    pays = df_p[df_p['Invoice_Ref'].isin(invs['Invoice_No'].astype(str))]

    total_quote = float(qs['Total_Value'].sum())
    total_billed = float(invs['Split_Amount'].sum())
    total_collected = float(pays['Amount'].sum())

    # ✅ Lifecycle composition (no overlap)
    val_collected = total_collected
    val_outstanding = max(0.0, total_billed - total_collected)
    val_unbilled = max(0.0, total_quote - total_billed)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Project Value", f"{total_quote:,.0f}")
    c2.metric("Invoiced (Billed)", f"{total_billed:,.0f}")
    c3.metric("Collected", f"{total_collected:,.0f}")
    c4.metric("Outstanding", f"{val_outstanding:,.0f}", delta=float(-val_outstanding), delta_color="inverse")

    st.divider()

    c_pie1, c_pie2 = st.columns([1, 2])
    with c_pie1:
        st.write("#### 🌍 Value Composition (Lifecycle)")
        if total_quote > 0:
            df_global = pd.DataFrame({
                'Category': ['Collected', 'Outstanding (Billed)', 'Unbilled Scope'],
                'Value': [val_collected, val_outstanding, val_unbilled]
            })
            fig_g = px.pie(df_global, values='Value', names='Category', hole=0.4)
            fig_g.update_layout(showlegend=True, height=300, margin=dict(l=10, r=10, t=10, b=10))
            st.plotly_chart(fig_g, use_container_width=True)
        else:
            st.info("No data available.")

    with c_pie2:
        st.write("#### 🏗️ Per-Project Composition (Lifecycle)")
        if not qs.empty:
            cols = st.columns(2)
            for idx, (_, q) in enumerate(qs.iterrows()):
                with cols[idx % 2]:
                    qid = str(q['Quote_ID'])
                    q_val = float(q['Total_Value'])

                    q_invs = invs[invs['Quote_Ref'].astype(str) == qid]
                    q_billed = float(q_invs['Split_Amount'].sum())

                    # payments allocated to this quote across its invoices
                    q_coll = 0.0
                    for inv_no in q_invs['Invoice_No'].astype(str).unique().tolist():
                        q_coll += float(_payments_for_invoice_quote(df_p, inv_no, qid)['Amount'].sum())

                    p_coll = q_coll
                    p_out = max(0.0, q_billed - q_coll)
                    p_unbill = max(0.0, q_val - q_billed)

                    # ✅ FIX: never overwrite df_p
                    df_pie = pd.DataFrame({'Type': ['Collected', 'Outstanding', 'Unbilled'], 'Val': [p_coll, p_out, p_unbill]})
                    fig_p = px.pie(df_pie, values='Val', names='Type', hole=0.4, title=f"<b>{q['Project_Name']}</b>")
                    fig_p.update_layout(showlegend=False, height=200, margin=dict(l=10, r=10, t=30, b=10))
                    st.plotly_chart(fig_p, use_container_width=True)

    st.divider()

    st.write("### ✅ Compliance Matrix")
    if qs.empty:
        st.info("No active projects.")
    else:
        for _, q in qs.iterrows():
            qid = str(q['Quote_ID'])
            q_invs = invs[invs['Quote_Ref'].astype(str) == qid]
            q_inv_ids = q_invs['Invoice_No'].astype(str).unique().tolist()

            q_pays = df_p[df_p['Invoice_Ref'].isin(q_inv_ids)].copy()
            # if allocated, show only this quote
            if q_pays['Quote_Ref'].astype(str).str.len().sum() > 0:
                q_pays = q_pays[q_pays['Quote_Ref'].astype(str) == qid]

            with st.expander(f"📂 {q['Project_Name']} (ID: {qid}) - {len(q_pays)} Payment Lines"):
                if not q_pays.empty:
                    data = []
                    for _, row in q_pays.iterrows():
                        data.append({
                            'Invoice': row['Invoice_Ref'],
                            'Quote': row.get('Quote_Ref', ''),
                            'Payment ID': row['Payment_ID'],
                            'Amount': f"{row['Amount']:,.2f}",
                            'Bank Slip': "✅" if str(row['Proof_File']) != "None" else "❌",
                            'Form C': "✅" if str(row['Form_C_File']) != "None" else "❌",
                            'Declaration': "✅" if str(row['Payment_Decl_File']) != "None" else "❌"
                        })
                    st.table(pd.DataFrame(data))
                else:
                    st.write("No payments yet.")


# --- TAB 1: QUOTATIONS ---
with tab1:
    st.write("### 📄 Contract Management")
    c_upl, c_prev = st.columns([1, 1])
    q_file = c_upl.file_uploader("Upload Agreement PDF", type=['pdf'], key='q_up')

    if q_file:
        with c_prev:
            with st.expander("📄 PDF Preview", expanded=True):
                display_pdf(q_file)

    if 'detected_sows' not in st.session_state:
        st.session_state.detected_sows = []

    if q_file and not st.session_state.detected_sows:
        with st.spinner("🤖 Scanning for 'Scope of Work' (SOW)..."):
            with open("temp_agree.pdf", "wb") as f:
                f.write(q_file.getbuffer())
            results = parse_multi_sow_agreement("temp_agree.pdf")
            if results:
                st.session_state.detected_sows = results
                st.success(f"✅ Found {len(results)} SOWs!")
            else:
                st.warning("No SOWs found. Use Manual Entry.")

    if st.session_state.detected_sows:
        with st.form("sow_form"):
            st.subheader("Review Detected Items")
            final_qs = []
            for idx, item in enumerate(st.session_state.detected_sows):
                c1, c2, c3 = st.columns([2, 1, 1])
                def_id = f"QT-{datetime.now().strftime('%y%m')}-{idx+1}"
                n = c1.text_input(f"Project Name #{idx+1}", value=item['name'])
                v = c2.number_input(f"Value #{idx+1}", value=float(item['amount']))
                i = c3.text_input(f"ID #{idx+1}", value=def_id)
                final_qs.append({'id': i, 'name': n, 'val': v, 'date': item['date']})

            if st.form_submit_button("💾 Save Detected Quotations"):
                for q in final_qs:
                    base = os.path.join(VAULT, curr_biz, q['id'], "Agreements")
                    fname = safe_copy(q_file, base, q_file.name)
                    new_row = {
                        'Quote_ID': q['id'], 'Date': q['date'], 'Business': curr_biz,
                        'Project_Name': q['name'], 'Total_Value': q['val'],
                        'Agreement_File': fname, 'Status': 'Open'
                    }
                    df_q = pd.concat([df_q, pd.DataFrame([new_row])], ignore_index=True)
                save_db(df_q, df_i, df_p, curr_biz)
                st.success("✅ Quotations Created!")
                st.session_state.detected_sows = []
                time.sleep(1)
                st.rerun()

    st.divider()
    with st.expander("➕ Add Manual Quotation"):
        with st.form("man_q"):
            c1, c2 = st.columns(2)
            mid = c1.text_input("Quote ID")
            mname = c2.text_input("Project Name")
            c3, c4 = st.columns(2)
            mval = c3.number_input("Value", min_value=0.0)
            mfile = c4.file_uploader("Agreement PDF (Optional)", type=['pdf'])

            if st.form_submit_button("Save Manual Quote"):
                if mid and mname:
                    fname = safe_copy(mfile, os.path.join(VAULT, curr_biz, mid, "Agreements"), mfile.name) if mfile else "None"
                    new_row = {
                        'Quote_ID': mid, 'Date': datetime.today(), 'Business': curr_biz,
                        'Project_Name': mname, 'Total_Value': mval, 'Agreement_File': fname, 'Status': 'Manual'
                    }
                    df_q = pd.concat([df_q, pd.DataFrame([new_row])], ignore_index=True)
                    save_db(df_q, df_i, df_p, curr_biz)
                    st.success("Saved!")
                    st.rerun()

    st.dataframe(df_q[df_q['Business'] == curr_biz])


# --- TAB 2: INVOICES ---
with tab2:
    st.write("### 🧾 Invoice Processing")

    col_up, col_view = st.columns([1, 1])
    i_file = col_up.file_uploader("Upload Invoice PDF", type=['pdf'], key='inv_up')
    i_decl_file = col_up.file_uploader("Upload Declaration/Mushak PDF", type=['pdf'], key='inv_dec')

    if i_file:
        with col_view:
            with st.expander("📄 Invoice Preview", expanded=True):
                display_pdf(i_file)

    if 'map_items' not in st.session_state:
        st.session_state.map_items = []

    if i_file and not st.session_state.map_items:
        with st.spinner("🔍 Analyzing Table Structure..."):
            with open("temp_inv.pdf", "wb") as f:
                f.write(i_file.getbuffer())
            data = parse_invoice_v2("temp_inv.pdf")
            st.session_state.inv_meta = data
            st.session_state.map_items = []
            for idx, item in enumerate(data['items']):
                st.session_state.map_items.append({
                    'id': idx,
                    'desc': item['desc'],
                    'amt': float(item['amount']),
                    'action': 'Existing Quote',
                    'target': '',
                    'alloc_amt': float(item['amount'])  # ✅ adjustable
                })

    if st.session_state.map_items:
        meta = st.session_state.inv_meta
        st.info(f"**Invoice:** {meta['no']} | **PDF Total:** {meta['total']:,.2f}")

        raw_opts = df_q[df_q['Business'] == curr_biz]
        smart_opts = raw_opts.apply(lambda x: f"{x['Quote_ID']} | {x['Project_Name']}", axis=1).tolist()

        st.write("👇 **Map Line Items to Quotations (with adjustable allocation amount)**")
        final_rows = []
        valid_form = True

        for item in st.session_state.map_items:
            c1, c2, c3, c4 = st.columns([2.2, 1.2, 2.3, 1.3])

            c1.write(f"**{item['desc'][:55]}...**")
            c1.caption(f"Detected amount: {item['amt']:,.2f}")

            act = c2.radio("Action", ["Existing Quote", "New Quote", "Ignore"], key=f"a_{item['id']}")
            target_id = ""
            new_q_data = None

            # ✅ NEW: adjustable allocation amount (even for existing quote)
            alloc_amt = c4.number_input(
                "Allocate",
                min_value=0.0,
                value=float(item.get('alloc_amt', item['amt'])),
                step=1.0,
                key=f"alloc_{item['id']}",
                help="How much of this invoice line will be posted against the selected quotation."
            )

            if act == "Existing Quote":
                sel_str = c3.selectbox("Select Quote", smart_opts, key=f"s_{item['id']}")
                if sel_str:
                    target_id = sel_str.split(' | ')[0]

            elif act == "New Quote":
                new_id = c3.text_input("New ID", value=f"QT-{meta['no']}-{item['id']}", key=f"ni_{item['id']}")
                new_nm = c3.text_input("New Project Name", value=item['desc'][:30], key=f"nn_{item['id']}")
                new_val = c3.number_input("New Quote Value", value=float(item['amt']), key=f"nv_{item['id']}")
                if new_id and new_nm:
                    new_q_data = {'id': new_id, 'name': new_nm, 'val': new_val}
                    target_id = new_id
                else:
                    valid_form = False

            # Basic validation: cannot allocate more than detected amount (soft rule)
            if alloc_amt > float(item['amt']) + 0.0001:
                valid_form = False
                st.warning(f"⚠️ Line #{item['id']+1}: allocation cannot exceed detected amount.")

            final_rows.append({'meta': item, 'target': target_id, 'new_q': new_q_data, 'action': act, 'alloc_amt': float(alloc_amt)})

        st.divider()
        if st.button("💾 Process Invoice"):
            if not i_file:
                st.error("⚠️ Invoice PDF missing!")
            elif not valid_form:
                st.error("⚠️ Fix mapping/allocation issues.")
            else:
                fname_inv = safe_copy(i_file, os.path.join(VAULT, curr_biz, "Invoices"), i_file.name)
                fname_dec = safe_copy(i_decl_file, os.path.join(VAULT, curr_biz, "Invoices"), i_decl_file.name) if i_decl_file else "None"

                count = 0
                for r in final_rows:
                    if r['action'] == "Ignore":
                        continue

                    if r['new_q']:
                        d = r['new_q']
                        os.makedirs(os.path.join(VAULT, curr_biz, d['id'], "Invoices"), exist_ok=True)
                        q_row = {
                            'Quote_ID': d['id'], 'Date': meta['date'], 'Business': curr_biz,
                            'Project_Name': d['name'], 'Total_Value': d['val'], 'Status': 'Auto',
                            'Agreement_File': "None"
                        }
                        df_q = pd.concat([df_q, pd.DataFrame([q_row])], ignore_index=True)

                    inv_row = {
                        'Invoice_No': str(meta['no']),
                        'Quote_Ref': str(r['target']),
                        'Date': meta['date'],
                        'Business': curr_biz,
                        'Split_Amount': float(r['alloc_amt']),  # ✅ allocated amount saved
                        'Description': r['meta']['desc'],
                        'Invoice_File': fname_inv,
                        'Declaration_File': fname_dec
                    }
                    df_i = pd.concat([df_i, pd.DataFrame([inv_row])], ignore_index=True)
                    count += 1

                save_db(df_q, df_i, df_p, curr_biz)
                st.success(f"✅ Saved {count} allocated invoice line(s)!")
                st.session_state.map_items = []
                time.sleep(1)
                st.rerun()


# --- TAB 3: PAYMENTS ---
with tab3:
    st.write("### 💵 Payment Collection (supports partial + allocation per quotation)")

    curr_invs = df_i[df_i['Business'] == curr_biz].copy()
    if curr_invs.empty:
        st.info("No invoices for this business yet.")
    else:
        # invoice totals (all quote lines under that invoice)
        inv_totals = curr_invs.groupby('Invoice_No')['Split_Amount'].sum()

        # payment totals by invoice (sum of all allocation lines)
        pay_totals = df_p.groupby('Invoice_Ref')['Amount'].sum()

        unpaid_list = []
        for inv_no, val in inv_totals.items():
            paid = float(pay_totals.get(str(inv_no), 0.0))
            rem = float(val) - paid
            if rem > 1.0:
                unpaid_list.append(f"{inv_no} (Due: {rem:,.0f})")

        st.subheader("1. Record New Payment")
        if not unpaid_list:
            st.success("🎉 All Invoices Paid!")
        else:
            sel_str = st.selectbox("Select Invoice to Pay", unpaid_list)
            sel_inv_no = sel_str.split(" ")[0]

            due_val = float(inv_totals[sel_inv_no]) - float(pay_totals.get(sel_inv_no, 0.0))
            due_val = max(0.0, due_val)

            # ✅ Determine the quote lines under this invoice (invoice may map to multiple quotations)
            inv_lines = curr_invs[curr_invs['Invoice_No'].astype(str) == str(sel_inv_no)].copy()
            inv_lines['Quote_Ref'] = inv_lines['Quote_Ref'].astype(str)

            # For each quote inside this invoice, compute due = billed - paid(allocated)
            quote_due_rows = []
            for qref, grp in inv_lines.groupby('Quote_Ref'):
                billed_q = float(grp['Split_Amount'].sum())
                paid_q = float(df_p[(df_p['Invoice_Ref'].astype(str) == str(sel_inv_no)) & (df_p['Quote_Ref'].astype(str) == str(qref))]['Amount'].sum())
                due_q = max(0.0, billed_q - paid_q)
                quote_due_rows.append({'Quote_Ref': qref, 'Billed': billed_q, 'Paid': paid_q, 'Due': due_q})

            quote_due_df = pd.DataFrame(quote_due_rows).sort_values('Quote_Ref')
            multi_quote = len(quote_due_df) > 1

            with st.form("pay_form"):
                c1, c2 = st.columns(2)
                p_amt = c1.number_input("Amount Received (partial allowed)", min_value=0.0, value=float(due_val), step=1.0)
                p_date = c2.date_input("Date", value=datetime.today())

                st.write("📎 **Attachments (At least one required)**")
                c_f1, c_f2, c_f3 = st.columns(3)
                f_proof = c_f1.file_uploader("🏦 Bank Slip/Proof", type=['pdf', 'jpg'])
                f_formc = c_f2.file_uploader("📄 Form C (Optional)", type=['pdf'])
                f_decl = c_f3.file_uploader("📝 Declaration (Optional)", type=['pdf'])

                if f_proof and getattr(f_proof, "type", "") == "application/pdf":
                    with st.expander("📄 Proof Preview"):
                        display_pdf(f_proof)

                st.divider()
                st.write("### 🔁 Allocate this payment to quotation(s) under the selected invoice")

                alloc_inputs = {}
                if quote_due_df.empty:
                    st.warning("No invoice lines found to allocate. (Unexpected)")
                else:
                    st.dataframe(quote_due_df, use_container_width=True, height=180)

                    if multi_quote:
                        st.caption("This invoice is linked to multiple quotations. Allocate the received amount across quotations below.")
                    else:
                        st.caption("This invoice is linked to a single quotation. Allocation will be auto-filled (you can still change).")

                    for _, r in quote_due_df.iterrows():
                        qref = str(r['Quote_Ref'])
                        default_alloc = min(float(r['Due']), float(p_amt)) if not multi_quote else 0.0
                        alloc_inputs[qref] = st.number_input(
                            f"Allocate to Quote {qref}",
                            min_value=0.0,
                            value=float(default_alloc),
                            step=1.0,
                            key=f"alloc_pay_{sel_inv_no}_{qref}"
                        )

                    # If single quote and user didn't change, push full payment
                    if not multi_quote and quote_due_df.shape[0] == 1:
                        only_q = str(quote_due_df.iloc[0]['Quote_Ref'])
                        # make it easier: show hint
                        st.caption(f"Tip: set allocation for {only_q} to exactly {p_amt:,.0f} to match received amount.")

                if st.form_submit_button("💾 Record Payment"):
                    if not (f_proof or f_formc or f_decl):
                        st.error("⚠️ You must upload at least one attachment.")
                    else:
                        # ✅ Validate allocation sums
                        alloc_sum = float(sum(alloc_inputs.values())) if alloc_inputs else 0.0
                        if abs(alloc_sum - float(p_amt)) > 0.0001:
                            st.error(f"⚠️ Allocation total must equal Amount Received. Received={p_amt:,.2f} but Allocated={alloc_sum:,.2f}")
                        else:
                            # Save attachments once (per invoice/first quote folder)
                            # choose a stable folder: invoice payments under business -> InvoiceNo
                            save_path = os.path.join(VAULT, curr_biz, "Payments", str(sel_inv_no))
                            os.makedirs(save_path, exist_ok=True)

                            n_proof = safe_copy(f_proof, save_path, f_proof.name) if f_proof else "None"
                            n_formc = safe_copy(f_formc, save_path, f_formc.name) if f_formc else "None"
                            n_decl = safe_copy(f_decl, save_path, f_decl.name) if f_decl else "None"

                            parent_id = f"PAY-{int(time.time())}"
                            line_no = 0

                            # ✅ Create one payment row per quote allocation (this enables perfect tracking)
                            for qref, amt in alloc_inputs.items():
                                if float(amt) <= 0:
                                    continue
                                line_no += 1
                                new_row = {
                                    'Payment_ID': f"{parent_id}-{line_no}",
                                    'Parent_Payment_ID': parent_id,
                                    'Invoice_Ref': str(sel_inv_no),
                                    'Quote_Ref': str(qref),
                                    'Date': p_date,
                                    'Amount': float(amt),
                                    'Proof_File': n_proof,
                                    'Form_C_File': n_formc,
                                    'Payment_Decl_File': n_decl
                                }
                                df_p = pd.concat([df_p, pd.DataFrame([new_row])], ignore_index=True)

                            save_db(df_q, df_i, df_p, curr_biz)
                            st.success("✅ Payment Recorded with allocation!")
                            time.sleep(1)
                            st.rerun()

        st.divider()
        st.subheader("2. Manage & Update Payments")

        # show payments relevant to this business invoices
        biz_inv_nums = curr_invs['Invoice_No'].astype(str).tolist()
        biz_pays = df_p[df_p['Invoice_Ref'].astype(str).isin(biz_inv_nums)].copy()

        if biz_pays.empty:
            st.info("No payments found.")
        else:
            biz_pays['display'] = biz_pays.apply(
                lambda x: f"{x['Payment_ID']} | Parent: {x.get('Parent_Payment_ID','')} | Inv: {x['Invoice_Ref']} | Quote: {x.get('Quote_Ref','')} | {float(x['Amount']):,.2f}",
                axis=1
            )
            pay_opts = biz_pays['display'].tolist()
            sel_pay_str = st.selectbox("Select Payment Line to View/Update", pay_opts)

            if sel_pay_str:
                sel_pay_id = sel_pay_str.split(' | ')[0]
                sel_row = df_p[df_p['Payment_ID'] == sel_pay_id].iloc[0]

                # files are stored under VAULT/business/Payments/InvoiceNo
                base_path = os.path.join(VAULT, curr_biz, "Payments", str(sel_row['Invoice_Ref']))

                c_info, c_prev = st.columns([1, 1])
                with c_info:
                    st.write("**Payment Details**")
                    st.write(f"- Invoice: `{sel_row['Invoice_Ref']}`")
                    st.write(f"- Quote: `{sel_row.get('Quote_Ref','')}`")
                    st.write(f"- Amount: `{float(sel_row['Amount']):,.2f}`")
                    st.write(f"- Parent: `{sel_row.get('Parent_Payment_ID','')}`")

                    st.write("**Current Attachments:**")
                    st.write(f"🏦 Bank Slip: `{sel_row['Proof_File']}`")
                    st.write(f"📄 Form C: `{sel_row['Form_C_File']}`")
                    st.write(f"📝 Declaration: `{sel_row['Payment_Decl_File']}`")

                    with st.form("update_doc_form"):
                        u_formc = st.file_uploader("Upload Form C", type=['pdf'])
                        u_decl = st.file_uploader("Upload Declaration", type=['pdf'])
                        if st.form_submit_button("💾 Update Documents"):
                            idx = df_p[df_p['Payment_ID'] == sel_pay_id].index[0]
                            if u_formc:
                                fname = safe_copy(u_formc, base_path, u_formc.name)
                                df_p.at[idx, 'Form_C_File'] = fname
                            if u_decl:
                                fname = safe_copy(u_decl, base_path, u_decl.name)
                                df_p.at[idx, 'Payment_Decl_File'] = fname
                            save_db(df_q, df_i, df_p, curr_biz)
                            st.success("✅ Documents Updated!")
                            time.sleep(1)
                            st.rerun()

                with c_prev:
                    st.write("**👁️ Preview Attachment**")
                    prev_opt = st.radio("Select File", ["Bank Slip", "Form C", "Declaration"], horizontal=True)
                    file_path = "None"
                    if prev_opt == "Bank Slip":
                        file_path = str(sel_row['Proof_File'])
                    elif prev_opt == "Form C":
                        file_path = str(sel_row['Form_C_File'])
                    elif prev_opt == "Declaration":
                        file_path = str(sel_row['Payment_Decl_File'])

                    if file_path != "None" and file_path != "nan":
                        full_p = os.path.join(base_path, file_path)
                        display_pdf(full_p)
                    else:
                        st.info("No file attached.")


# --- TAB 4: MASTER LEDGER ---
with tab4:
    st.write("### 📊 Financial Master Ledger")
    if st.button("🔄 Refresh & Export"):
        save_db(df_q, df_i, df_p, curr_biz)
        st.success(f"✅ Exported to {FILE}")

    view_df = generate_ledger_view(curr_biz, df_q, df_i, df_p)

    def style_df(row):
        bg = ''
        if row.get('Type') == 'QUOTE':
            bg = 'background-color: #E3F2FD; font-weight: bold'
        elif row.get('Type') == 'INVOICE':
            bg = 'background-color: #FFF9C4'
        elif row.get('Type') == 'PAYMENT':
            bg = 'background-color: #E8F5E9'
        elif row.get('Type') == 'SUMMARY':
            bg = 'background-color: #F5F5F5; font-weight: bold'
        elif row.get('Type') == 'GRAND':
            bg = 'background-color: #212121; color: white; font-weight: bold'

        s = [bg] * len(row)
        if row.get('Status', '') == '🔴':
            s[-1] += '; color: red; font-weight: bold'
        elif row.get('Status', '') in ('✅', '🟢'):
            s[-1] += '; color: green; font-weight: bold'
        return s

    st.dataframe(view_df.style.apply(style_df, axis=1), height=800, use_container_width=True)


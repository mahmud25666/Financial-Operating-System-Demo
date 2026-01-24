import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import os
import time
import pdfplumber
import re

# --- 1. CONFIGURATION ---
st.set_page_config(page_title="Glafit Empire Finance", layout="wide", page_icon="🏢")

FILE = 'Finance_Ledger.xlsx'
VAULT_FOLDER = 'Master_Vault'

if not os.path.exists(VAULT_FOLDER):
    os.makedirs(VAULT_FOLDER)

# --- 2. EMBEDDED PARSER (No need for ingest.py) ---
def parse_invoice(file_path):
    """
    Reads a PDF and extracts Invoice No, Date, Amount, and Project.
    Embedded directly here to avoid ImportErrors.
    """
    try:
        with pdfplumber.open(file_path) as pdf:
            text = ""
            for page in pdf.pages:
                text += page.extract_text() or ""
        
        # Simple Regex Logic (Adjust as needed)
        # 1. Invoice Number (Looks for INV-...)
        inv_match = re.search(r'(INV-\d{4}-\d{3}|INV-\w+-\d+)', text)
        invoice_no = inv_match.group(0) if inv_match else f"INV-{int(time.time())}"
        
        # 2. Amount (Looks for $ or numbers with decimals)
        amt_match = re.search(r'\$\s?([\d,]+\.\d{2})', text)
        if amt_match:
            amount = float(amt_match.group(1).replace(',', ''))
        else:
            # Fallback: look for largest number
            nums = re.findall(r'([\d,]+\.\d{2})', text)
            amount = float(max(nums).replace(',', '')) if nums else 0.0
            
        # 3. Date
        date_match = re.search(r'(\d{4}-\d{2}-\d{2}|\d{2}/\d{2}/\d{4})', text)
        inv_date = pd.to_datetime(date_match.group(0)) if date_match else datetime.today()

        # 4. Project Name (Heuristic)
        project = "General Project"
        lines = text.split('\n')
        for line in lines[:10]: # Check first 10 lines
            if "Project:" in line:
                project = line.replace("Project:", "").strip()
                break
                
        return invoice_no, inv_date, amount, project

    except Exception as e:
        return "MANUAL_CHECK", datetime.today(), 0.0, "Manual Entry"

# --- 3. DATA HELPER FUNCTIONS ---

def ensure_columns_exist(df, required_cols):
    for col in required_cols:
        if col not in df.columns:
            df[col] = "" 
    return df

def get_data():
    try:
        df_inv = pd.read_excel(FILE, sheet_name='Invoices')
        df_pay = pd.read_excel(FILE, sheet_name='Payments')
        
        df_inv = ensure_columns_exist(df_inv, ['Project_Name', 'Business_Unit', 'Client', 'Invoice_No', 'Total_Amount', 'Date', 'Entry_Date', 'PDF_File'])
        df_pay = ensure_columns_exist(df_pay, ['Invoice_Ref', 'Amount_Received', 'Payment_Date', 'Entry_Date', 'Proof_File'])
        
        if not df_pay.empty:
            df_pay['Date'] = pd.to_datetime(df_pay['Payment_Date'], errors='coerce')
            df_pay['Entry_Date'] = pd.to_datetime(df_pay['Entry_Date'], errors='coerce')
        else:
            df_pay['Date'] = pd.to_datetime([])

        if not df_inv.empty:
            df_inv['Date'] = pd.to_datetime(df_inv['Date'], errors='coerce')
            df_inv['Entry_Date'] = pd.to_datetime(df_inv['Entry_Date'], errors='coerce')

        df_inv['Invoice_No'] = df_inv['Invoice_No'].astype(str)
        df_inv['Business_Unit'] = df_inv['Business_Unit'].astype(str)
        
        return df_inv, df_pay
    except Exception as e:
        return pd.DataFrame(), pd.DataFrame()

def sync_ledger_to_excel(df_inv, df_pay):
    """
    Creates a SEPARATE Ledger Sheet for EACH Business Unit.
    Matches the Dashboard coloring and structure.
    """
    from openpyxl.styles import PatternFill
    
    if df_inv.empty: return
    all_businesses = df_inv['Business_Unit'].unique().tolist()
    
    with pd.ExcelWriter(FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        for biz_name in all_businesses:
            biz_inv = df_inv[df_inv['Business_Unit'] == biz_name].copy()
            if biz_inv.empty: continue

            biz_inv_nums = biz_inv['Invoice_No'].unique()
            biz_pay = df_pay[df_pay['Invoice_Ref'].isin(biz_inv_nums)].copy() if not df_pay.empty else pd.DataFrame()

            ledger_rows = []
            cumulative_balance = 0.0
            
            sorted_inv = biz_inv.sort_values(by='Date')

            for idx, inv in sorted_inv.iterrows():
                try: inv_amt = float(inv['Total_Amount'])
                except: inv_amt = 0.0
                cumulative_balance += inv_amt
                
                t_date = inv['Date'].strftime('%Y-%m-%d') if pd.notnull(inv['Date']) else ""
                e_date = inv['Entry_Date'].strftime('%Y-%m-%d %H:%M') if pd.notnull(inv['Entry_Date']) else "Legacy"

                ledger_rows.append({
                    'Transaction_Date': t_date, 'System_Entry_Date': e_date,
                    'Description': f"INVOICE: {inv['Invoice_No']} ({inv['Project_Name']})",
                    'Debit': inv_amt, 'Credit': 0.0, 'Balance': cumulative_balance, 'Type': 'Invoice'
                })

                pay_total_for_this_inv = 0.0

                if not biz_pay.empty:
                    my_pays = biz_pay[biz_pay['Invoice_Ref'] == inv['Invoice_No']].sort_values(by='Date')
                    for p_idx, pay in my_pays.iterrows():
                        try: p_amt = float(pay['Amount_Received'])
                        except: p_amt = 0.0
                        cumulative_balance -= p_amt
                        pay_total_for_this_inv += p_amt
                        
                        pt_date = pay['Date'].strftime('%Y-%m-%d') if pd.notnull(pay['Date']) else ""
                        pe_date = pay['Entry_Date'].strftime('%Y-%m-%d %H:%M') if pd.notnull(pay['Entry_Date']) else "Legacy"

                        ledger_rows.append({
                            'Transaction_Date': pt_date, 'System_Entry_Date': pe_date,
                            'Description': f"   >>> Payment Received",
                            'Debit': 0.0, 'Credit': p_amt, 'Balance': cumulative_balance, 'Type': 'Payment'
                        })

                remaining = inv_amt - pay_total_for_this_inv
                ledger_rows.append({
                    'Transaction_Date': "", 'System_Entry_Date': "",
                    'Description': f"   >>> Remaining Due for {inv['Invoice_No']}",
                    'Debit': None, 'Credit': None, 'Balance': remaining, 'Type': 'Summary'
                })
                ledger_rows.append({'Type': 'Spacer'})

            ledger_rows.append({
                'Transaction_Date': "TOTALS", 'System_Entry_Date': datetime.now().strftime('%Y-%m-%d %H:%M'),
                'Description': "💰 GRAND TOTAL OUTSTANDING",
                'Debit': sum([r['Debit'] for r in ledger_rows if r.get('Type')=='Invoice']),
                'Credit': sum([r['Credit'] for r in ledger_rows if r.get('Type')=='Payment']),
                'Balance': cumulative_balance, 'Type': 'GrandTotal'
            })

            sheet_name = f"Ldg-{str(biz_name)[:26]}"
            df_export = pd.DataFrame(ledger_rows)
            cols = ['Transaction_Date', 'System_Entry_Date', 'Description', 'Debit', 'Credit', 'Balance', 'Type']
            for c in cols: 
                if c not in df_export.columns: df_export[c] = ""
            df_export = df_export[cols]
            df_export.to_excel(writer, sheet_name=sheet_name, index=False)

            # --- COLORING ---
            ws = writer.book[sheet_name]
            yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            green = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            red = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            paid_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

            for row in ws.iter_rows(min_row=2, max_col=7):
                cell_type = row[6].value
                cell_bal = row[5].value
                fill = None
                if cell_type == 'Invoice': fill = yellow
                elif cell_type == 'Payment': fill = green
                elif cell_type in ['Summary', 'GrandTotal']:
                    try: val = float(cell_bal)
                    except: val = 0.0
                    fill = red if val > 0.01 else paid_fill
                if fill:
                    for cell in row: cell.fill = fill

df_inv, df_pay = get_data()

# --- 3. SIDEBAR ---
st.sidebar.title("🏢 Business Units")

def get_all_businesses():
    if os.path.exists(VAULT_FOLDER):
        folder_biz = [d for d in os.listdir(VAULT_FOLDER) if os.path.isdir(os.path.join(VAULT_FOLDER, d))]
    else:
        folder_biz = []
    
    if not df_inv.empty:
        excel_biz = [x for x in df_inv['Business_Unit'].unique().tolist() if str(x).lower() != 'nan']
    else:
        excel_biz = []
    combined = sorted(list(set(folder_biz + excel_biz)))
    return combined if combined else ["Main Business"]

existing_businesses = get_all_businesses()
business_selection = st.sidebar.selectbox("Select Dashboard:", existing_businesses + ["+ Add New Business"])
current_business = business_selection

if business_selection == "+ Add New Business":
    new_biz_name = st.sidebar.text_input("Enter New Business Name:")
    if st.sidebar.button("Create Business"):
        if new_biz_name:
            new_folder_path = os.path.join(VAULT_FOLDER, new_biz_name)
            if not os.path.exists(new_folder_path):
                os.makedirs(new_folder_path)
                st.sidebar.success(f"✅ Created folder for '{new_biz_name}'!")
                time.sleep(1); st.rerun()
            else:
                st.sidebar.warning("Business folder already exists.")

st.sidebar.divider()

# --- DATE RANGE FILTER ---
st.sidebar.subheader("📅 Date Filter")
min_date = df_inv['Date'].min() if not df_inv.empty and not pd.isnull(df_inv['Date'].min()) else datetime(2025, 1, 1)
max_date = datetime.today()

c1, c2 = st.sidebar.columns(2)
start_date = c1.date_input("From", min_date)
end_date = c2.date_input("To", max_date)

start_date = pd.to_datetime(start_date)
end_date = pd.to_datetime(end_date)

# --- 4. DATA LOGIC (VIEW GENERATOR) ---
if business_selection == "+ Add New Business" and current_business not in existing_businesses:
    filtered_inv = pd.DataFrame(columns=df_inv.columns)
else:
    filtered_inv = df_inv[df_inv['Business_Unit'] == current_business].copy()

if not df_pay.empty and not filtered_inv.empty:
    valid_inv_numbers = filtered_inv['Invoice_No'].unique()
    filtered_pay = df_pay[df_pay['Invoice_Ref'].isin(valid_inv_numbers)].copy()
else:
    filtered_pay = pd.DataFrame(columns=df_pay.columns)

ledger_rows = []
cumulative_balance = 0.0

if not filtered_inv.empty:
    filtered_inv = filtered_inv.sort_values(by='Date')
    
    for idx, inv in filtered_inv.iterrows():
        if not (start_date <= inv['Date'] <= end_date): continue

        try: inv_amount = float(inv['Total_Amount'])
        except: inv_amount = 0.0
        cumulative_balance += inv_amount
        
        inv_file = str(inv.get('PDF_File', '')).strip()
        inv_link = os.path.join(os.getcwd(), VAULT_FOLDER, current_business, inv_file) if inv_file and inv_file != "Manual_Entry" else None

        ledger_rows.append({
            'Date': inv['Date'],
            'Description': f"🟦 INVOICE: {inv['Invoice_No']} ({inv['Project_Name']})",
            'Project': inv['Project_Name'], 
            'Debit': inv_amount, 'Credit': 0.0, 'Balance': cumulative_balance,
            'Link_Path': inv_link, 'Type': 'Invoice'
        })

        pay_total_for_this_inv = 0.0

        if not filtered_pay.empty:
            my_payments = filtered_pay[filtered_pay['Invoice_Ref'] == inv['Invoice_No']].sort_values(by='Date')
            for p_idx, pay in my_payments.iterrows():
                try: pay_amount = float(pay['Amount_Received'])
                except: pay_amount = 0.0
                cumulative_balance -= pay_amount
                pay_total_for_this_inv += pay_amount
                
                pay_file = str(pay.get('Proof_File', '')).strip()
                pay_link = os.path.join(os.getcwd(), VAULT_FOLDER, current_business, "Payments", pay_file) if pay_file and pay_file != "Manual_Entry" else None
                e_date = pay.get('Entry_Date')
                entry_str = e_date.strftime('%Y-%m-%d') if pd.notnull(e_date) else "Legacy"

                ledger_rows.append({
                    'Date': pay['Date'], 
                    'Description': f"   ↘ 🟩 Payment (Entry: {entry_str})",
                    'Project': inv['Project_Name'], 
                    'Debit': 0.0, 'Credit': pay_amount, 'Balance': cumulative_balance,
                    'Link_Path': pay_link, 'Type': 'Payment'
                })

        remaining = inv_amount - pay_total_for_this_inv
        status_icon = "✅" if remaining < 0.01 else "⏳"
        ledger_rows.append({
            'Date': None,
            'Description': f"   👉 {status_icon} Remaining Due for {inv['Invoice_No']}", 
            'Project': None,
            'Debit': None, 'Credit': None, 'Balance': remaining,
            'Link_Path': None, 'Type': 'Summary'
        })

df_view = pd.DataFrame(ledger_rows)

if not df_view.empty:
    df_view.loc[len(df_view)] = {
        'Date': None, 'Description': "<b>💰 TOTAL CUMULATIVE OUTSTANDING</b>", 'Project': None,
        'Debit': df_view[df_view['Type'] == 'Invoice']['Debit'].sum(),
        'Credit': df_view[df_view['Type'] == 'Payment']['Credit'].sum(),
        'Balance': cumulative_balance, 'Link_Path': None, 'Type': 'GrandTotal'
    }

# --- 5. DASHBOARD UI (ENHANCED) ---
st.title(f"📊 Dashboard: {current_business}")

if not df_view.empty:
    metrics_df = df_view[df_view['Type'].isin(['Invoice', 'Payment'])].copy()
    total_billed = metrics_df['Debit'].sum()
    total_paid = metrics_df['Credit'].sum()
    current_balance = df_view.iloc[-1]['Balance']
    collection_rate = (total_paid / total_billed * 100) if total_billed > 0 else 0
else:
    metrics_df = pd.DataFrame()
    total_billed = 0.0; total_paid = 0.0; current_balance = 0.0; collection_rate = 0.0

# ROW 1: KPIs
k1, k2, k3, k4 = st.columns(4)
k1.metric("Total Billed", f"${total_billed:,.2f}")
k2.metric("Total Collected", f"${total_paid:,.2f}")
k3.metric("Outstanding Due", f"${current_balance:,.2f}", delta="Balance", delta_color="inverse")
k4.metric("Collection Rate", f"{collection_rate:.1f}%")

st.divider()

if not metrics_df.empty:
    c1, c2 = st.columns([2, 1])
    
    with c1:
        st.subheader("📈 Financial Velocity (Volume)")
        chart_data = metrics_df.copy()
        chart_data['Cumulative Billed'] = chart_data[chart_data['Type']=='Invoice']['Debit'].cumsum()
        chart_data['Cumulative Collected'] = chart_data[chart_data['Type']=='Payment']['Credit'].cumsum()
        
        chart_data = chart_data.set_index('Date').sort_index()
        chart_data['Cumulative Billed'] = chart_data['Cumulative Billed'].ffill().fillna(0)
        chart_data['Cumulative Collected'] = chart_data['Cumulative Collected'].ffill().fillna(0)
        chart_data = chart_data.reset_index()

        fig_trend = px.area(chart_data, x='Date', y=['Cumulative Billed', 'Cumulative Collected'], 
                            title="Accumulated Revenue & Collections",
                            color_discrete_map={'Cumulative Billed': '#EF553B', 'Cumulative Collected': '#00CC96'})
        fig_trend.update_layout(hovermode="x unified", template="plotly_white", yaxis_title="Amount ($)")
        st.plotly_chart(fig_trend, use_container_width=True)
        
    with c2:
        st.subheader("💳 Revenue Status")
        visual_outstanding = max(0, current_balance) 
        visual_collected = total_paid
        if current_balance < 0: visual_collected = total_billed + abs(current_balance) 

        pie_data = pd.DataFrame({'Status': ['Collected', 'Outstanding'], 'Amount': [visual_collected, visual_outstanding]})
        fig_pie = px.pie(pie_data, names='Status', values='Amount', hole=0.6, 
                         color='Status', color_discrete_map={'Collected':'#00CC96', 'Outstanding':'#EF553B'})
        
        center_text = "Overpaid" if current_balance < 0 else f"{collection_rate:.0f}%"
        fig_pie.update_layout(annotations=[dict(text=center_text, x=0.5, y=0.5, font_size=20, showarrow=False)])
        st.plotly_chart(fig_pie, use_container_width=True)

    c3, c4 = st.columns([2, 1])

    with c3:
        st.subheader("📅 Net Monthly Cash Flow")
        cf_df = metrics_df.copy()
        cf_df['Month'] = cf_df['Date'].dt.strftime('%Y-%m')
        
        monthly = cf_df.groupby(['Month', 'Type'])[['Debit', 'Credit']].sum().reset_index()
        monthly['Amount'] = monthly.apply(lambda x: x['Debit'] if x['Type'] == 'Invoice' else x['Credit'], axis=1)
        monthly['Category'] = monthly['Type'].map({'Invoice': 'Invoiced', 'Payment': 'Collected'})
        
        net_flow = cf_df.groupby('Month').apply(lambda x: x[x['Type']=='Payment']['Credit'].sum() - x[x['Type']=='Invoice']['Debit'].sum()).reset_index(name='Net Cash Flow')

        fig_combo = go.Figure()
        for cat, color in [('Invoiced', '#EF553B'), ('Collected', '#00CC96')]:
            subset = monthly[monthly['Category'] == cat]
            fig_combo.add_trace(go.Bar(x=subset['Month'], y=subset['Amount'], name=cat, marker_color=color))
        
        fig_combo.add_trace(go.Scatter(x=net_flow['Month'], y=net_flow['Net Cash Flow'], mode='lines+markers', 
                                       name='Net Monthly Cash', line=dict(color='blue', width=3)))
        
        fig_combo.update_layout(barmode='group', title="Monthly Invoiced vs. Collected + Net Flow", template="plotly_white")
        st.plotly_chart(fig_combo, use_container_width=True)

    with c4:
        st.subheader("🏆 Top Projects")
        proj_df = metrics_df[metrics_df['Type'] == 'Invoice'].groupby('Project')['Debit'].sum().reset_index()
        proj_df = proj_df.sort_values(by='Debit', ascending=True).tail(5) 
        
        fig_proj = px.bar(proj_df, x='Debit', y='Project', orientation='h', 
                          title="Top Revenue Drivers", text_auto='.2s',
                          color_discrete_sequence=['#636EFA'])
        fig_proj.update_layout(xaxis_title="Total Billed ($)", yaxis_title="")
        st.plotly_chart(fig_proj, use_container_width=True)

else:
    st.info("Start adding invoices and payments to see the analytics graphs!")

# --- 6. ACTION TABS ---
tab1, tab2, tab3 = st.tabs(["📝 Add Invoice", "💵 Record Payment", "📄 Master Ledger"])

with tab1:
    st.subheader(f"Upload Invoice to: {current_business}")
    uploaded_file = st.file_uploader("Drag & Drop PDF Invoice", type=['pdf'])
    
    def_inv, def_date, def_amt, def_proj = "", datetime.today(), 0.0, ""
    
    if uploaded_file:
        biz_folder = os.path.join(VAULT_FOLDER, current_business)
        if not os.path.exists(biz_folder): os.makedirs(biz_folder)
        
        temp_path = os.path.join(biz_folder, uploaded_file.name)
        with open(temp_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        
        st.success(f"File saved to: {current_business}/{uploaded_file.name}")
        
        with st.spinner("🤖 AI reading invoice..."):
            parsed_inv, parsed_date, parsed_amt, parsed_proj = parse_invoice(temp_path)
            def_inv = parsed_inv if parsed_inv != "MANUAL_CHECK" else ""
            def_amt = parsed_amt
            def_proj = parsed_proj
            if parsed_date: def_date = parsed_date 
            
            existing_inv_nums = df_inv['Invoice_No'].astype(str).str.strip().tolist()
            if str(def_inv).strip() in existing_inv_nums:
                st.warning(f"⚠️ Warning: Invoice '{def_inv}' already exists.")

    with st.form("new_invoice"):
        c1, c2 = st.columns(2)
        new_inv = c1.text_input("Invoice Number", value=def_inv)
        new_date = c2.date_input("Date", value=def_date)
        c3, c4 = st.columns(2)
        new_amt = c3.number_input("Total Amount ($)", min_value=0.0, value=float(def_amt))
        new_proj = c4.text_input("Project Name", value=def_proj)
        
        if st.form_submit_button("💾 Save Invoice"):
            current_invoices = df_inv['Invoice_No'].astype(str).str.strip().tolist()
            if str(new_inv).strip() in current_invoices:
                st.error(f"❌ STOP: Invoice '{new_inv}' already exists!")
            elif new_inv and new_amt > 0:
                new_row = {
                    'Invoice_No': new_inv, 'Date': new_date, 'Client': 'Manual Client',
                    'Project_Name': new_proj, 'Total_Amount': new_amt,
                    'PDF_File': uploaded_file.name if uploaded_file else "Manual_Entry",
                    'Business_Unit': current_business,
                    'Entry_Date': datetime.now()
                }
                
                df_new = pd.DataFrame([new_row])
                correct_cols = ['Invoice_No', 'Date', 'Entry_Date', 'Client', 'Project_Name', 'Total_Amount', 'PDF_File', 'Business_Unit']
                for c in correct_cols:
                     if c not in df_new.columns: df_new[c] = ""
                df_new = df_new[correct_cols]
                
                with pd.ExcelWriter(FILE, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df_new.to_excel(writer, sheet_name='Invoices', index=False, header=False, startrow=len(df_inv)+1)
                
                df_inv_new, df_pay_new = get_data()
                sync_ledger_to_excel(df_inv_new, df_pay_new)
                st.success("Invoice Saved & Excel Ledger Synced!"); time.sleep(1); st.rerun()
            else: 
                st.error("Missing Info!")

with tab2:
    st.subheader("Record Payment")
    
    unpaid_map = {} 
    if not filtered_inv.empty:
        for idx, row in filtered_inv.iterrows():
            inv_no = str(row['Invoice_No'])
            try: total = float(row['Total_Amount'])
            except: total = 0.0
            if not filtered_pay.empty:
                paid = filtered_pay[filtered_pay['Invoice_Ref'] == inv_no]['Amount_Received'].sum()
            else: paid = 0.0
            remaining = total - paid
            if remaining > 0.01: unpaid_map[inv_no] = remaining
    
    unpaid_list = list(unpaid_map.keys())

    if unpaid_list:
        with st.form("pay_form"):
            c1, c2 = st.columns(2)
            p_inv = c1.selectbox("Select Invoice to Pay:", unpaid_list)
            curr_due = unpaid_map[p_inv]
            c1.caption(f"Current Due: ${curr_due:,.2f}")
            
            p_date = c2.date_input("Payment Date (Check/Bank Date)", datetime.today())
            p_amt = c1.number_input("Amount Received ($)", value=float(curr_due))
            pay_file = c2.file_uploader("Attach Proof (Optional)", type=['pdf', 'png', 'jpg'])
            
            if st.form_submit_button("Confirm Payment"):
                proof_filename = "Manual_Entry"
                if pay_file:
                    pay_folder = os.path.join(VAULT_FOLDER, current_business, "Payments")
                    if not os.path.exists(pay_folder): os.makedirs(pay_folder)
                    save_path = os.path.join(pay_folder, pay_file.name)
                    with open(save_path, "wb") as f: f.write(pay_file.getbuffer())
                    proof_filename = pay_file.name
                
                new_pay = {
                    'Payment_ID': f"PAY-{datetime.now().strftime('%H%M%S')}",
                    'Invoice_Ref': p_inv, 'Amount_Received': p_amt, 'Method': 'Manual',
                    'Proof_File': proof_filename, 'Payment_Date': p_date,
                    'Entry_Date': datetime.now()
                }

                df_new_pay = pd.DataFrame([new_pay])
                correct_pay_cols = ['Payment_ID', 'Invoice_Ref', 'Amount_Received', 'Method', 'Proof_File', 'Payment_Date', 'Entry_Date']
                for c in correct_pay_cols:
                     if c not in df_new_pay.columns: df_new_pay[c] = ""
                df_new_pay = df_new_pay[correct_pay_cols]
                
                with pd.ExcelWriter(FILE, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df_new_pay.to_excel(writer, sheet_name='Payments', index=False, header=False, startrow=len(df_pay)+1)
                
                df_inv_new, df_pay_new = get_data()
                sync_ledger_to_excel(df_inv_new, df_pay_new)
                st.success("Payment Recorded & Excel Ledger Synced!"); time.sleep(1); st.rerun()
    else:
        st.info("All invoices are fully paid! 🎉")

with tab3:
    st.subheader("Detailed Ledger (Grouped by Invoice)")
    if not df_view.empty:
        display_df = df_view.copy()
        display_df['Date'] = pd.to_datetime(display_df['Date'], errors='coerce').dt.strftime('%Y-%m-%d').fillna("")
        
        def highlight_ledger(row):
            bg = ''
            if row['Type'] == 'Invoice': bg = 'background-color: #FFF9C4' 
            elif row['Type'] == 'Payment': bg = 'background-color: #E8F5E9' 
            elif row['Type'] in ['GrandTotal', 'Summary']:
                if row['Balance'] > 0.01: bg = 'background-color: #FFCDD2' 
                else: bg = 'background-color: #C8E6C9' 
            return [bg] * len(row)

        styled_df = display_df.style.apply(highlight_ledger, axis=1)

        st.dataframe(
            styled_df,
            column_order=['Date', 'Description', 'Debit', 'Credit', 'Balance', 'Link_Path'],
            column_config={
                "Debit": st.column_config.NumberColumn(format="$%.2f"),
                "Credit": st.column_config.NumberColumn(format="$%.2f"),
                "Balance": st.column_config.NumberColumn(format="$%.2f"),
                "Link_Path": st.column_config.LinkColumn("Attachment", display_text="Open File", help="Click to open"),
            },
            use_container_width=True, height=700, hide_index=True
        )
        st.caption("ℹ️ Color Legend: 🟨 Invoice | 🟩 Payment | 🟥 Outstanding Balance | 🟢 Fully Paid")
        st.caption("ℹ️ Backup Note: A full backup of this view is automatically saved to the 'Master_Ledger' sheet in your Excel file.")
    else:
        st.info("No transactions found.")
